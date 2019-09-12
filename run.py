import openpyxl
import csv
#csvfile = "input.csv"
csvfile = "/Users/hegderajesh/Downloads/MUHURTHA REGISTRATION.csv"
xlfile = 'format.xlsx'

# initializing the titles and rows list 
fields = [] 
rows = []

# reading csv file 
with open(csvfile, 'r') as csvfile: 
    # creating a csv reader object 
    csvreader = csv.reader(csvfile) 
      
    # extracting field names through first row 
    fields = next(csvreader) 
  
    # extracting each data row one by one 
    for row in csvreader: 
        rows.append(row)

 # open xlsx file and select sheet1
wb = openpyxl.load_workbook(filename=xlfile)
ws = wb.get_sheet_by_name('Sheet1')

for i in range(len(rows)):
    print('Processing', i ,'row')
    # writing data to cell in loop
    ws.cell(row=5, column=4, value=rows[i][3])
    ws.cell(row=5, column=3, value='Name of '+rows[i][2])
    ws.cell(row=5, column=7, value=rows[i][4])
    ws.cell(row=6, column=4, value=rows[i][5])
    ws.cell(row=6, column=7, value=rows[i][6])
    ws.cell(row=7, column=4, value=rows[i][7])
    ws.cell(row=7, column=7, value=rows[i][8])
    ws.cell(row=8, column=4, value=rows[i][9])
    ws.cell(row=8, column=7, value=rows[i][10])
    ws.cell(row=9, column=4, value=rows[i][11])
    ws.cell(row=9, column=7, value=rows[i][12])
    ws.cell(row=10, column=4, value=rows[i][13])
    ws.cell(row=10, column=7, value=rows[i][14])
    ws.cell(row=11, column=4, value=rows[i][15])
    ws.cell(row=11, column=7, value=rows[i][16])
    ws.cell(row=12, column=4, value=rows[i][17])
    ws.cell(row=12, column=7, value=rows[i][18])
    ws.cell(row=13, column=4, value=rows[i][19])
    ws.cell(row=13, column=7, value=rows[i][20])
    ws.cell(row=14, column=4, value=rows[i][1])
    ws.cell(row=14, column=7, value=rows[i][21])
    ws.cell(row=16, column=7, value=rows[i][0])
    ws.cell(row=21, column=3, value=rows[i][25])
    ws.cell(row=23, column=4, value=rows[i][23])
    # save file
    new_filename = rows[i][1]
    wb.save('/Users/hegderajesh/Documents/Website/Muhurtha/MuhurthaFiles/'+new_filename+'.xlsx')
    print('Done. Saved!')